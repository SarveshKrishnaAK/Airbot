"""
Excel Generation Service for Aerospace Test Cases
Generates professionally styled Excel files from test case data
"""

import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle
)
from openpyxl.utils import get_column_letter
from loguru import logger


class ExcelService:
    # Color scheme matching the frontend
    COLORS = {
        'header_bg': '0D47A1',      # Deep blue
        'header_fg': 'FFFFFF',      # White text
        'subheader_bg': '1565C0',   # Lighter blue
        'row_alt': 'E3F2FD',        # Light blue for alternating rows
        'success': '4CAF50',        # Green for pass
        'warning': 'FF9800',        # Orange for pending
        'error': 'F44336',          # Red for fail
        'critical': 'D32F2F',       # Dark red for critical
        'border': 'BBDEFB',         # Light border
    }

    def __init__(self):
        logger.info("Excel Service initialized")

    def parse_test_case_content(self, content: str) -> list:
        """Parse the LLM response to extract test case fields"""
        test_cases = []

        # Split by TEST CASE markers
        tc_pattern = r'\*\*TEST CASE\*\*[\s\S]*?(?=\*\*TEST CASE\*\*|$)'
        matches = re.findall(tc_pattern, content, re.IGNORECASE)

        if not matches:
            # Try to parse as single test case
            matches = [content]

        for tc_content in matches:
            test_case = self._extract_fields(tc_content)
            if test_case.get('id') or test_case.get('title'):
                test_cases.append(test_case)

        return test_cases if test_cases else [self._extract_fields(content)]

    def _extract_fields(self, content: str) -> dict:
        """Extract individual fields from test case content"""
        def extract_field(pattern, default=''):
            match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
            return match.group(1).strip() if match else default

        def extract_list(pattern):
            match = re.search(pattern, content, re.IGNORECASE | re.DOTALL)
            if not match:
                return []
            list_content = match.group(1)
            items = re.findall(r'[-*\d.]+\s*(.+?)(?=\n[-*\d.]|\n\*\*|$)', list_content)
            return [item.strip() for item in items if item.strip()]

        return {
            'id': extract_field(r'\*\*ID:\*\*\s*(.+?)(?:\n|$)', 'TC-001'),
            'title': extract_field(r'\*\*Title:\*\*\s*(.+?)(?:\n|$)', 'Test Case'),
            'system': extract_field(r'\*\*System Under Test:\*\*\s*(.+?)(?:\n|$)', '-'),
            'standards': extract_field(r'\*\*Applicable Standards:\*\*\s*(.+?)(?:\n|$)', '-'),
            'description': extract_field(r'\*\*Description:\*\*\s*([\s\S]+?)(?=\n\*\*Preconditions|$)', '-'),
            'preconditions': extract_list(r'\*\*Preconditions:\*\*([\s\S]*?)(?=\*\*Test Equipment|\*\*Test Steps|$)'),
            'equipment': extract_list(r'\*\*Test Equipment Required:\*\*([\s\S]*?)(?=\*\*Test Steps|$)'),
            'steps': extract_list(r'\*\*Test Steps:\*\*([\s\S]*?)(?=\*\*Expected|$)'),
            'expected': extract_list(r'\*\*Expected Results:\*\*([\s\S]*?)(?=\*\*Failure|$)'),
            'failure_criteria': extract_list(r'\*\*Failure Criteria:\*\*([\s\S]*?)(?=\*\*Actual|$)'),
            'actual': extract_field(r'\*\*Actual Results:\*\*\s*(.+?)(?:\n\*\*|$)', 'To be filled'),
            'status': extract_field(r'\*\*Status:\*\*\s*(.+?)(?:\n|$)', 'PENDING'),
            'priority': extract_field(r'\*\*Priority:\*\*\s*(.+?)(?:\n|$)', 'MEDIUM'),
            'category': extract_field(r'\*\*Category:\*\*\s*(.+?)(?:\n|$)', 'Functional'),
            'duration': extract_field(r'\*\*Estimated Duration:\*\*\s*(.+?)(?:\n|$)', '-'),
            'risk': extract_field(r'\*\*Risk Level:\*\*\s*(.+?)(?:\n|$)', 'Medium'),
        }

    def generate_excel(self, content: str, query: str = "") -> io.BytesIO:
        """Generate a professionally styled Excel file from test case content"""
        logger.info("Generating Excel file...")

        test_cases = self.parse_test_case_content(content)
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Define styles
        header_font = Font(bold=True, color=self.COLORS['header_fg'], size=12)
        header_fill = PatternFill(start_color=self.COLORS['header_bg'],
                                   end_color=self.COLORS['header_bg'],
                                   fill_type='solid')
        subheader_fill = PatternFill(start_color=self.COLORS['subheader_bg'],
                                      end_color=self.COLORS['subheader_bg'],
                                      fill_type='solid')
        alt_fill = PatternFill(start_color=self.COLORS['row_alt'],
                                end_color=self.COLORS['row_alt'],
                                fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='thin', color=self.COLORS['border']),
            bottom=Side(style='thin', color=self.COLORS['border'])
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Title row
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"AIRBOT - Aerospace Test Cases Report"
        title_cell.font = Font(bold=True, size=16, color=self.COLORS['header_fg'])
        title_cell.fill = header_fill
        title_cell.alignment = center_alignment
        ws.row_dimensions[1].height = 30

        # Metadata row
        ws.merge_cells('A2:N2')
        meta_cell = ws['A2']
        meta_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Query: {query[:100]}{'...' if len(query) > 100 else ''}"
        meta_cell.font = Font(italic=True, size=10)
        meta_cell.fill = alt_fill
        meta_cell.alignment = center_alignment

        # Headers
        headers = [
            'ID', 'Title', 'System Under Test', 'Standards', 'Description',
            'Preconditions', 'Test Equipment', 'Test Steps', 'Expected Results',
            'Failure Criteria', 'Actual Results', 'Status', 'Priority', 'Category'
        ]

        header_row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = subheader_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[header_row].height = 25

        # Data rows
        for idx, tc in enumerate(test_cases):
            row = header_row + 1 + idx
            is_alt = idx % 2 == 1

            data = [
                tc.get('id', ''),
                tc.get('title', ''),
                tc.get('system', ''),
                tc.get('standards', ''),
                tc.get('description', ''),
                '\n'.join(f"• {item}" for item in tc.get('preconditions', [])) or '-',
                '\n'.join(f"• {item}" for item in tc.get('equipment', [])) or '-',
                '\n'.join(f"{i+1}. {step}" for i, step in enumerate(tc.get('steps', []))) or '-',
                '\n'.join(f"• {item}" for item in tc.get('expected', [])) or '-',
                '\n'.join(f"• {item}" for item in tc.get('failure_criteria', [])) or '-',
                tc.get('actual', ''),
                tc.get('status', ''),
                tc.get('priority', ''),
                tc.get('category', ''),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment

                if is_alt:
                    cell.fill = alt_fill

                # Color-code status
                if col == 12:  # Status column
                    status = str(value).upper()
                    if 'PASS' in status:
                        cell.font = Font(bold=True, color=self.COLORS['success'])
                    elif 'FAIL' in status:
                        cell.font = Font(bold=True, color=self.COLORS['error'])
                    else:
                        cell.font = Font(bold=True, color=self.COLORS['warning'])

                # Color-code priority
                if col == 13:  # Priority column
                    priority = str(value).upper()
                    if 'CRITICAL' in priority:
                        cell.font = Font(bold=True, color=self.COLORS['critical'])
                    elif 'HIGH' in priority:
                        cell.font = Font(bold=True, color=self.COLORS['error'])

            # Set row height for content
            ws.row_dimensions[row].height = 80

        # Column widths
        col_widths = [12, 25, 20, 20, 35, 30, 25, 40, 35, 30, 20, 12, 15, 15]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        # Freeze panes
        ws.freeze_panes = 'A5'

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Excel file generated with {len(test_cases)} test cases")
        return output


excel_service = ExcelService()
